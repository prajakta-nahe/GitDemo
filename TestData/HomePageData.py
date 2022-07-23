import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname":"Prajakta", "email":"abc@gmail.com", "gender":"Female"}, {"firstname":"Pranav", "email":"xyz@gmail.com", "gender":"Male"}]

    @staticmethod     # Declaring as a static method so as to avoid creating object , and call the method directly by class name
    def getTestData(test_case_name):    # Removing self parameter as it applies to only non static method
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Lenovo\\Documents\\PythonDemo.xlsx")  # to provide the path of excel to be opened #use double slash
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # To get rows
            if sheet.cell(row=i, column=1).value == test_case_name :  # add this step to print only Testcase2 row
                for j in range(2, sheet.max_column + 1):  # starting from column 2
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value  # To get only firstname

        return[Dict]             # Returning as a list for multiple data sets