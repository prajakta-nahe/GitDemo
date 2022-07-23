import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Lenovo\\Documents\\PythonDemo.xlsx")   # to provide the path of excel to be opened #use double slash
sheet = book.active                                                             # to select opened active sheet
Dict ={}
cell = sheet.cell(row=1, column=2)                                              # to select the cell to be printed
print(cell.value)                                                               # cell value
sheet.cell(row=2,column=2).value = "Prajakta"                                   # To write into the cell, locate the cell
print(sheet.cell(row=2,column=2).value)                                         # Print the value in console
book.save("C:\\Users\\Lenovo\\Documents\\PythonDemo.xlsx")                      # Need to save the excel, only then the value will be printed in excel
                                                                    # Note - The excel should be closed while running the program so as to write into it
print(sheet.max_row)                                                # To find maximum number of rows in sheet
print(sheet.max_column)
print(sheet["B2"].value)
                                # to print every value present in the sheet
for i in range(1,sheet.max_row+1):                                  # To get rows
    if sheet.cell(row = i,column =1).value == "Testcase2":             # add this step to print only Testcase2 row
        #for j in range(1,sheet.max_column+1):                         # To get columns
            #print(sheet.cell(row=i, column=j).value)
        for j in range(2, sheet.max_column + 1):                                        #starting from column 2
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value         # To get only firstname
print(Dict)

